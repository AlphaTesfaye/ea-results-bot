"""
excel_export.py
----------------
Builds an in-memory .xlsx file for an announcement's candidate list,
so the full list can be attached to a Telegram message instead of
being truncated in the message text.
"""

import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def _safe_sheet_name(name: str) -> str:
    """Excel sheet names: max 31 chars, no []:*?/\\"""
    cleaned = re.sub(r'[\[\]:*?/\\]', "", name or "Candidates")
    return (cleaned[:31] or "Candidates")


def build_candidate_xlsx(ann) -> io.BytesIO:
    """Returns an in-memory xlsx file (BytesIO) for the given announcement's candidates."""
    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_name(ann.position)

    # Header info rows
    ws.append(["Position", ann.position])
    ws.append(["Location", ann.location])
    ws.append(["Announcement", ann.announcement_type])
    ws.append([])  # blank spacer row

    if ann.candidates:
        columns = list(ann.candidates[0].keys())
        header_row_num = ws.max_row + 1
        ws.append(columns)
        for cell in ws[header_row_num]:
            cell.font = Font(bold=True)

        for row in ann.candidates:
            ws.append([row.get(col, "") for col in columns])

        # Auto-width columns (rough heuristic based on content length)
        for i, col in enumerate(columns, 1):
            max_len = max(
                [len(str(col))] + [len(str(row.get(col, ""))) for row in ann.candidates]
            )
            ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 50)

    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="left")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def safe_filename(ann) -> str:
    """Generate a filesystem/Telegram-safe filename for the attachment."""
    base = re.sub(r'[^A-Za-z0-9 _-]', "", ann.position or "candidates")
    base = re.sub(r'\s+', "_", base.strip()) or "candidates"
    return f"{base[:60]}.xlsx"
